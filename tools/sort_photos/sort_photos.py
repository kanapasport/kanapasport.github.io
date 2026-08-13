#!/usr/bin/env python3
"""
sort_photos.py (v3)
====================
Trideni fotek z pasportizace budov do slozek podle mistnosti a technologie.

Cilova struktura:
    <dest>/<BUDOVA>/<MISTNOST>/<TECHNOLOGIE>/fotky...
napr.
    G:/pasport/G62/GS642/VZT/IMG_0123.jpg

Rezimy (--mode):
    sort   (vychozi) - roztridi fotky
    learn            - nauci se nazvy mistnosti z rucne opravenych slozek
    undo             - vrati zpet cely predchozi beh podle logu

--------------------------------------------------------------------------
JAK TO FUNGUJE
--------------------------------------------------------------------------
1. FAZE ANALYZY (paralelni, vyuziva vsechna jadra procesoru)
   Kazda fotka projde rychlym levnym testem "je na ni vubec text?".
   Vetsina fotek (vzduchotechnika, hasici pristroje, vybaveni) tim skonci
   a dal se nezpracovava. Jen fotky, ktere testem projdou jako kandidat na
   "fotku tabletu s PDF", jdou do plne (drahe) analyzy:
     - detekce a orez displeje tabletu (pokud je nainstalovan OpenCV)
     - OCR ve 4 variantach predzpracovani (kvuli sedemu, hure citelnemu textu)
     - sparovani nalezeneho kodu se seznamem skutecnych mistnosti (fuzzy match)

2. FAZE PRIRAZENI (sekvencni, v chronologickem poradi)
   Fotky se prochazi v poradi poriz., a plati: fotka tabletu otevre novou
   mistnost, vsechny nasledujici fotky patri do ni, dokud neprijde dalsi
   fotka tabletu.

3. FAZE ZAPISU + KONTROLY
   Presun/kopie souboru, kontrola integrity (nic se neztratilo), zapis
   logu pro pripadne vraceni zpet, a Excel report.

--------------------------------------------------------------------------
INSTALACE (jednou na kazdem pocitaci)
--------------------------------------------------------------------------
  1) Python 3.11+  https://www.python.org/downloads/
     (pri instalaci ZASKRTNOUT "Add python.exe to PATH")
  2) Tesseract OCR https://github.com/UB-Mannheim/tesseract/wiki
     Pri instalaci pridat cesky jazyk. Zapamatovat cestu, typicky:
       C:\\Program Files\\Tesseract-OCR\\tesseract.exe
  3) pip install -r requirements.txt

--------------------------------------------------------------------------
PRIKLADY SPUSTENI
--------------------------------------------------------------------------
Nanecisto (nic se nepresune, jen se vypise co by se stalo + Excel report):
  python sort_photos.py --inbox "D:\\Inbox" --dest "D:\\Pasport" ^
      --tesseract "C:\\Program Files\\Tesseract-OCR\\tesseract.exe" --dry-run

Ostry beh (presune fotky, zapise log pro undo):
  python sort_photos.py --inbox "D:\\Inbox" --dest "D:\\Pasport" ^
      --tesseract "C:\\Program Files\\Tesseract-OCR\\tesseract.exe" --move

Naucit se mistnosti z rucne opravenych slozek:
  python sort_photos.py --mode learn --dest "D:\\Pasport"

Vratit zpet posledni beh:
  python sort_photos.py --mode undo --dest "D:\\Pasport" --run-id 20260806_143000
"""

import argparse
import csv
import datetime as dt
import difflib
import hashlib
import json
import os
import re
import shutil
import sys
import time
from concurrent.futures import ProcessPoolExecutor
from dataclasses import dataclass, field
from pathlib import Path

# --------------------------------------------------------------------------
# Zavislosti
# --------------------------------------------------------------------------

try:
    from PIL import Image, ImageOps, ImageFilter
except ImportError:
    print("CHYBI knihovna Pillow.  Spustte:  pip install pillow")
    sys.exit(1)

try:
    import pytesseract
except ImportError:
    print("CHYBI knihovna pytesseract.  Spustte:  pip install pytesseract")
    sys.exit(1)

# Volitelne - HEIC fotky z iPhonu
try:
    import pillow_heif
    pillow_heif.register_heif_opener()
    HEIC_OK = True
except Exception:
    HEIC_OK = False

# Volitelne - detekce a orez displeje tabletu
try:
    import cv2
    import numpy as np
    CV2_OK = True
except Exception:
    CV2_OK = False

# Volitelne - Excel report
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    XLSX_OK = True
except Exception:
    XLSX_OK = False


# --------------------------------------------------------------------------
# Konstanty a nastaveni
# --------------------------------------------------------------------------

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".heic", ".heif", ".tif", ".tiff"}

DEFAULT_ROOM_PATTERN = r"\b[A-Z]{1,4}[-\s]?\d{2,4}[A-Z]?\b"

# Kolik znaku textu musi najit rychly test, aby se fotka poslala do plne analyzy.
MIN_TEXT_LEN_FOR_TABLET = 15
# Kolikrat se musi kod mistnosti objevit napric OCR variantami, aby byl "jisty".
MIN_CONFIDENCE_HITS = 2
# Skore fuzzy shody, nad kterym povazujeme sparovani se seznamem za spolehlive.
STRONG_MATCH_SCORE = 0.95

# Ocekavany pocet fotek na jednu mistnost - mimo tento rozsah = anomalie k rucni kontrole.
EXPECTED_SERIES_MIN = 3
EXPECTED_SERIES_MAX = 25

# Nazvy sluzebnich souboru v cilove slozce
LEDGER_FILE = "_zpracovane_fotky.json"       # evidence jiz zpracovanych fotek (idempotence)
VISION_CACHE_FILE = "_vision_cache.json"     # cache placenych vision volani
RUNS_DIR = "_behy"                           # logy jednotlivych behu (pro undo)
MASTER_XLSX = "_PREHLED_pasportizace.xlsx"   # souhrnny Excel pres vsechny behy

VISION_MODEL = "claude-sonnet-5"

UNKNOWN_BUILDING = "_neznama_budova"
UNKNOWN_ROOM_PREFIX = "Neznama_mistnost"
BEFORE_FIRST_TABLET = "_pred_prvnim_tabletem"


# ==========================================================================
# KONFIGURACE BUDOV A TECHNOLOGII
# ==========================================================================

@dataclass
class BuildingConfig:
    """Konfigurace jedne budovy: kod budovy + seznam vsech platnych mistnosti."""
    building_code: str
    rooms: list
    room_pattern_str: str = DEFAULT_ROOM_PATTERN
    fuzzy_threshold: float = 0.8
    source_path: Path = None

    @classmethod
    def load(cls, path: Path) -> "BuildingConfig":
        data = json.loads(path.read_text(encoding="utf-8"))
        return cls(
            building_code=data.get("building_code", path.stem).upper(),
            rooms=[str(r).upper().strip() for r in data.get("rooms", []) if str(r).strip()],
            room_pattern_str=data.get("room_code_pattern", DEFAULT_ROOM_PATTERN),
            fuzzy_threshold=float(data.get("fuzzy_match_threshold", 0.8)),
            source_path=path,
        )

    def save_rooms(self, new_rooms: list):
        """Prida nove mistnosti do JSON konfigurace (se zalohou puvodniho souboru)."""
        data = json.loads(self.source_path.read_text(encoding="utf-8"))
        existing = {str(r).upper().strip() for r in data.get("rooms", [])}
        added = [r for r in new_rooms if r.upper() not in existing]
        if not added:
            return []
        backup = self.source_path.with_suffix(".json.bak")
        shutil.copy2(self.source_path, backup)
        data["rooms"] = sorted(existing | {a.upper() for a in added})
        self.source_path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        self.rooms = data["rooms"]
        return added


@dataclass
class RoomIndex:
    """Sloucena databaze mistnosti pres VSECHNY budovy.

    Diky tomu nemusi uzivatel pri spusteni vybirat budovu - kod mistnosti
    sam urci, do ktere budovy fotka patri (napr. G61 vs G62)."""
    room_to_building: dict = field(default_factory=dict)
    all_rooms: list = field(default_factory=list)
    buildings: dict = field(default_factory=dict)   # kod budovy -> BuildingConfig
    pattern: re.Pattern = None
    fuzzy_threshold: float = 0.8

    @classmethod
    def load_dir(cls, configs_dir: Path) -> "RoomIndex":
        idx = cls()
        patterns = []
        thresholds = []
        cfg_files = sorted(
            p for p in configs_dir.glob("*.json")
            if not p.name.endswith(".example.json") and not p.name.endswith(".bak")
        )
        for cfg_file in cfg_files:
            try:
                cfg = BuildingConfig.load(cfg_file)
            except Exception as e:
                print(f"  [!] Nelze nacist konfiguraci {cfg_file.name}: {e}")
                continue
            idx.buildings[cfg.building_code] = cfg
            patterns.append(cfg.room_pattern_str)
            thresholds.append(cfg.fuzzy_threshold)
            for room in cfg.rooms:
                if room in idx.room_to_building and idx.room_to_building[room] != cfg.building_code:
                    print(f"  [!] Mistnost {room} je uvedena u vice budov "
                          f"({idx.room_to_building[room]} i {cfg.building_code}) - "
                          f"pouzije se {idx.room_to_building[room]}.")
                    continue
                idx.room_to_building[room] = cfg.building_code
                idx.all_rooms.append(room)

        # Vsechny vzory spojime do jednoho (staci, aby sedel kterykoliv).
        if patterns:
            uniq = list(dict.fromkeys(patterns))
            combined = "|".join(f"(?:{p})" for p in uniq)
        else:
            combined = DEFAULT_ROOM_PATTERN
        idx.pattern = re.compile(combined, re.IGNORECASE)
        idx.fuzzy_threshold = min(thresholds) if thresholds else 0.8
        return idx

    def match(self, candidate: str):
        """Najde nejpodobnejsi znamou mistnost. Vraci (kod, budova, skore)."""
        if not candidate:
            return None, None, 0.0
        cand = candidate.upper()
        if cand in self.room_to_building:
            return cand, self.room_to_building[cand], 1.0
        if not self.all_rooms:
            return None, None, 0.0
        close = difflib.get_close_matches(cand, self.all_rooms, n=1, cutoff=0.0)
        if not close:
            return None, None, 0.0
        best = close[0]
        score = difflib.SequenceMatcher(None, cand, best).ratio()
        return best, self.room_to_building.get(best), score


def load_technologies(path: Path) -> dict:
    """Nacte prevod zkratek technologii na plne nazvy (jen pro report)."""
    if not path.exists():
        return {}
    try:
        return {k.upper(): v for k, v in json.loads(path.read_text(encoding="utf-8")).items()}
    except Exception:
        return {}


# ==========================================================================
# PRACE S OBRAZKY
# ==========================================================================

def open_image(path: Path) -> Image.Image:
    img = Image.open(path)
    img.load()
    return img


def get_exif_datetime(path: Path) -> str:
    """Vrati cas poriz. fotky jako 'YYYY:MM:DD HH:MM:SS', nebo '' kdyz chybi."""
    try:
        with Image.open(path) as img:
            exif = img.getexif()
            if exif:
                for tag_id in (36867, 306):   # DateTimeOriginal, DateTime
                    val = exif.get(tag_id)
                    if val:
                        return str(val).strip()
    except Exception:
        pass
    return ""


def detect_screen_crop(img: Image.Image):
    """Zkusi na fotce najit displej tabletu a vyriznout ho (vcetne narovnani
    perspektivy). Vraci orezany obrazek, nebo None kdyz se displej nenajde.

    Duvod: OCR na cele fotce se rozptyluje okolim (stul, ruka, pozadi mistnosti).
    Kdyz OCR dostane jen samotny displej, presnost cteni sedeho textu vyrazne roste.
    Vyzaduje OpenCV - bez nej se tento krok proste preskoci."""
    if not CV2_OK:
        return None
    try:
        arr = np.array(img.convert("RGB"))
        h0, w0 = arr.shape[:2]
        # Zmensime pro rychlost - hledani obrysu nepotrebuje plne rozliseni.
        scale = 1000.0 / max(h0, w0)
        if scale < 1.0:
            small = cv2.resize(arr, (int(w0 * scale), int(h0 * scale)))
        else:
            small = arr
            scale = 1.0

        gray = cv2.cvtColor(small, cv2.COLOR_RGB2GRAY)
        blurred = cv2.GaussianBlur(gray, (5, 5), 0)
        edges = cv2.Canny(blurred, 50, 150)
        edges = cv2.dilate(edges, np.ones((3, 3), np.uint8), iterations=1)

        contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if not contours:
            return None

        img_area = small.shape[0] * small.shape[1]
        best = None
        best_area = 0
        for cnt in sorted(contours, key=cv2.contourArea, reverse=True)[:10]:
            area = cv2.contourArea(cnt)
            # Displej musi zabirat rozumnou cast fotky, jinak to neni on.
            if area < img_area * 0.12 or area > img_area * 0.98:
                continue
            approx = cv2.approxPolyDP(cnt, 0.02 * cv2.arcLength(cnt, True), True)
            if len(approx) == 4 and cv2.isContourConvex(approx) and area > best_area:
                best = approx.reshape(4, 2).astype("float32")
                best_area = area
        if best is None:
            return None

        # Body zpet do plneho rozliseni
        best = best / scale

        # Serazeni rohu: levy-horni, pravy-horni, pravy-dolni, levy-dolni
        s = best.sum(axis=1)
        d = np.diff(best, axis=1).ravel()
        ordered = np.array([
            best[np.argmin(s)],   # levy horni
            best[np.argmin(d)],   # pravy horni
            best[np.argmax(s)],   # pravy dolni
            best[np.argmax(d)],   # levy dolni
        ], dtype="float32")

        (tl, tr, br, bl) = ordered
        width = int(max(np.linalg.norm(br - bl), np.linalg.norm(tr - tl)))
        height = int(max(np.linalg.norm(tr - br), np.linalg.norm(tl - bl)))
        if width < 200 or height < 200:
            return None

        dst = np.array([[0, 0], [width - 1, 0], [width - 1, height - 1], [0, height - 1]],
                       dtype="float32")
        matrix = cv2.getPerspectiveTransform(ordered, dst)
        warped = cv2.warpPerspective(arr, matrix, (width, height))
        return Image.fromarray(warped)
    except Exception:
        return None


def preprocess_variants(img: Image.Image):
    """Nekolik verzi obrazku pro OCR - resi sedy, malo kontrastni text na displeji."""
    gray = ImageOps.grayscale(img)
    autoc = ImageOps.autocontrast(gray, cutoff=2)
    variants = [gray, autoc]
    # Binarizace - vytahne slaby sedy text ze svetleho pozadi
    variants.append(autoc.point(lambda p: 255 if p > 150 else 0))
    # Zvetseni + doostreni - pomaha u maleho textu na fotce displeje
    w, h = gray.size
    if w * h < 12_000_000:      # u obrich fotek zvetsovani nema smysl
        upscaled = autoc.resize((w * 2, h * 2), Image.LANCZOS).filter(ImageFilter.SHARPEN)
        variants.append(upscaled)
    return variants


# ==========================================================================
# OCR
# ==========================================================================

def quick_text_probe(img: Image.Image, lang: str) -> str:
    """Rychly levny prvni pruchod - filtr. Vetsina fotek (vybaveni, technologie)
    nema zadny text a timhle se levne vyradi, aniz by na ne slo tezke OCR."""
    try:
        gray = ImageOps.grayscale(img)
        w, h = gray.size
        # Pro filtr staci zmenseny obrazek - jde jen o to zjistit, jestli tam JE text.
        if max(w, h) > 1600:
            ratio = 1600.0 / max(w, h)
            gray = gray.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
        return pytesseract.image_to_string(gray, lang=lang, config="--psm 6")
    except Exception:
        return ""


def full_ocr(img: Image.Image, lang: str) -> str:
    texts = []
    for variant in preprocess_variants(img):
        try:
            txt = pytesseract.image_to_string(variant, lang=lang)
        except Exception:
            txt = ""
        if txt.strip():
            texts.append(txt)
    return "\n---\n".join(texts)


def extract_room(text: str, pattern: re.Pattern, index: RoomIndex):
    """Z OCR textu vytahne nejpravdepodobnejsi kod mistnosti.

    Postup:
      1. regexem najde vsechny kandidaty
      2. kazdeho sparuje se seznamem skutecnych mistnosti (opravi OCR preklepy)
      3. z potvrzenych kandidatu vybere PROSTREDNI (podle poradi na strance) -
         hledany nazev mistnosti byva v nadpisu/uprostred planku, zatimco
         okolni kody jsou sousedni mistnosti uvedene jen orientacne

    Vraci (mistnost, budova, skore, pocet_vyskytu, seznam_kandidatu)."""
    # finditer + group(0) zamerne misto findall: kdyz vzor obsahuje zavorku
    # (napr. "(G6[12])?..."), findall vraci obsah zavorky misto cele shody,
    # takze by se misto "GS642" precetlo jen "G62".
    raw_matches = [m.group(0) for m in pattern.finditer(text)]
    raw_matches = [re.sub(r"[-\s]", "", m).upper() for m in raw_matches if m and m.strip()]
    if not raw_matches:
        return "", "", 0.0, 0, []

    confirmed = []      # (mistnost, budova, skore)
    for cand in raw_matches:
        room, building, score = index.match(cand)
        if room and score >= index.fuzzy_threshold:
            confirmed.append((room, building, score))

    pool = confirmed if confirmed else [(c, "", 0.0) for c in raw_matches]

    # Unikatni pri zachovani poradi vyskytu na strance
    seen = []
    for room, building, score in pool:
        if room not in [s[0] for s in seen]:
            seen.append((room, building, score))

    chosen_room, chosen_building, chosen_score = seen[len(seen) // 2]
    hits = sum(1 for room, _, _ in pool if room == chosen_room)
    return chosen_room, chosen_building or "", chosen_score, hits, raw_matches


# ==========================================================================
# PARALELNI ANALYZA FOTEK
# ==========================================================================

# Globalni stav pracovnich procesu (nastavi se pri jejich startu).
_WORKER = {}


def _worker_init(tesseract_cmd, lang, pattern_str, rooms, room_to_building,
                 fuzzy_threshold, use_screen_crop):
    if tesseract_cmd:
        pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
    try:
        import pillow_heif
        pillow_heif.register_heif_opener()
    except Exception:
        pass
    idx = RoomIndex(
        room_to_building=room_to_building,
        all_rooms=rooms,
        pattern=re.compile(pattern_str, re.IGNORECASE),
        fuzzy_threshold=fuzzy_threshold,
    )
    _WORKER["lang"] = lang
    _WORKER["index"] = idx
    _WORKER["pattern"] = idx.pattern
    _WORKER["crop"] = use_screen_crop


def _worker_analyze(path_str: str) -> dict:
    """Analyza jedne fotky - bezi v samostatnem procesu. Nic nemeni na disku."""
    path = Path(path_str)
    result = {
        "path": path_str,
        "exif": "",
        "is_tablet": False,
        "room": "",
        "building": "",
        "score": 0.0,
        "hits": 0,
        "candidates": [],
        "ocr_text": "",
        "screen_cropped": False,
        "error": "",
    }
    try:
        result["exif"] = get_exif_datetime(path)
        with open_image(path) as img:
            rgb = img.convert("RGB")

            probe = quick_text_probe(rgb, _WORKER["lang"])
            if len(probe.strip()) < MIN_TEXT_LEN_FOR_TABLET:
                # Fotka bez textu - temer jiste ne tablet. Konec, dal nezdrzujeme.
                result["ocr_text"] = probe.strip()[:200]
                return result

            target = rgb
            if _WORKER["crop"]:
                cropped = detect_screen_crop(rgb)
                if cropped is not None:
                    target = cropped
                    result["screen_cropped"] = True

            text = full_ocr(target, _WORKER["lang"])
            # Kdyz orez displeje nic nevydal, zkusime jeste celou fotku.
            if result["screen_cropped"] and len(text.strip()) < MIN_TEXT_LEN_FOR_TABLET:
                text = full_ocr(rgb, _WORKER["lang"])
                result["screen_cropped"] = False

            result["ocr_text"] = text[:500]
            room, building, score, hits, candidates = extract_room(
                text, _WORKER["pattern"], _WORKER["index"]
            )
            result["candidates"] = candidates
            if len(text.strip()) >= MIN_TEXT_LEN_FOR_TABLET and candidates:
                result["is_tablet"] = True
                result["room"] = room
                result["building"] = building
                result["score"] = score
                result["hits"] = hits
            elif len(text.strip()) >= MIN_TEXT_LEN_FOR_TABLET * 2:
                # Hodne textu, ale zadny kod mistnosti - nejspis tablet, jen necitelny.
                result["is_tablet"] = True
    except Exception as e:
        result["error"] = f"{type(e).__name__}: {e}"
    return result


# ==========================================================================
# VISION FALLBACK (volitelny, placeny)
# ==========================================================================

class VisionCache:
    def __init__(self, path: Path):
        self.path = path
        self.data = {}
        if path.exists():
            try:
                self.data = json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                self.data = {}

    def get(self, key):
        return self.data.get(key)

    def set(self, key, value):
        self.data[key] = value
        try:
            self.path.write_text(json.dumps(self.data, ensure_ascii=False, indent=2),
                                 encoding="utf-8")
        except Exception:
            pass


def call_vision(image_path: Path, rooms: list) -> str:
    """Necha fotku precist Claude vision modelem. Vraci kod mistnosti nebo ''.
    Vyzaduje:  pip install anthropic  +  promennou prostredi ANTHROPIC_API_KEY."""
    try:
        import anthropic
    except ImportError:
        print("  [!] Vision fallback vyzaduje:  pip install anthropic")
        return ""
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("  [!] Chybi promenna prostredi ANTHROPIC_API_KEY.")
        return ""

    import base64
    from io import BytesIO

    try:
        with open_image(image_path) as img:
            rgb = img.convert("RGB")
            crop = detect_screen_crop(rgb)
            target = crop if crop is not None else rgb
            # Zmenseni setri penize (min. tokenu) i cas, presnost to nezhorsi.
            target.thumbnail((1600, 1600), Image.LANCZOS)
            buf = BytesIO()
            target.save(buf, format="JPEG", quality=88)
            b64 = base64.standard_b64encode(buf.getvalue()).decode("ascii")
    except Exception as e:
        print(f"  [!] Nelze pripravit obrazek pro vision: {e}")
        return ""

    room_list = ", ".join(rooms[:400]) if rooms else "(seznam neni k dispozici)"
    prompt = (
        "Na fotce je tablet zobrazujici PDF planek mistnosti. V dokumentu je uveden "
        "kod/nazev mistnosti - text byva sedy a hure citelny.\n\n"
        "Pokud je na planku videt vic kodu mistnosti, vyber ten, ke kteremu se dokument "
        "primo vztahuje (nadpis stranky nebo mistnost zvyraznena v planku), nikoliv "
        "sousedni mistnosti uvedene jen orientacne.\n\n"
        f"Platne kody mistnosti pro tyto budovy: {room_list}\n\n"
        "Odpovez PRESNE jednim z techto tvaru, bez jakehokoliv dalsiho textu:\n"
        "ROOM: <kod mistnosti ze seznamu>\n"
        "UNKNOWN\n"
    )
    try:
        client = anthropic.Anthropic(api_key=api_key)
        resp = client.messages.create(
            model=VISION_MODEL,
            max_tokens=40,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image", "source": {"type": "base64",
                                                 "media_type": "image/jpeg", "data": b64}},
                    {"type": "text", "text": prompt},
                ],
            }],
        )
        text = resp.content[0].text.strip()
        if text.upper().startswith("ROOM:"):
            return text.split(":", 1)[1].strip().upper()
    except Exception as e:
        print(f"  [!] Vision API chyba: {e}")
    return ""


# ==========================================================================
# EVIDENCE ZPRACOVANYCH FOTEK (idempotence) A LOG PRO UNDO
# ==========================================================================

def quick_hash(path: Path) -> str:
    """Rychly otisk souboru: velikost + prvni 1 MB obsahu.

    Nehashuje se cely soubor zamerne - u desetitisicu fotek by to znamenalo
    precist desitky GB. Kombinace velikosti a zacatku souboru je pro rozpoznani
    "uz jsem tuhle fotku zpracoval" naprosto dostacujici."""
    h = hashlib.sha256()
    size = path.stat().st_size
    h.update(str(size).encode())
    with open(path, "rb") as f:
        h.update(f.read(1024 * 1024))
    return h.hexdigest()


class Ledger:
    """Evidence jiz zpracovanych fotek - aby opakovane spusteni na stejnou
    slozku nezpusobilo duplicity ani prehazeni jiz roztridenych fotek."""

    def __init__(self, path: Path):
        self.path = path
        self.data = {}
        if path.exists():
            try:
                self.data = json.loads(path.read_text(encoding="utf-8"))
            except Exception:
                self.data = {}

    def seen(self, file_hash: str) -> bool:
        return file_hash in self.data

    def record(self, file_hash: str, src: str, dst: str, run_id: str):
        self.data[file_hash] = {"src": src, "dst": dst, "run": run_id,
                                "cas": dt.datetime.now().isoformat(timespec="seconds")}

    def save(self):
        tmp = self.path.with_suffix(".tmp")
        tmp.write_text(json.dumps(self.data, ensure_ascii=False, indent=1), encoding="utf-8")
        tmp.replace(self.path)


def write_run_log(runs_dir: Path, run_id: str, moves: list, mode: str):
    """Zapise mapovani puvodni cesta -> nova cesta, aby sel beh vratit zpet."""
    runs_dir.mkdir(parents=True, exist_ok=True)
    log_path = runs_dir / f"beh_{run_id}.csv"
    with open(log_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["rezim", "zdroj", "cil"])
        for src, dst in moves:
            w.writerow([mode, src, dst])
    return log_path


def undo_run(dest: Path, run_id: str):
    """Vrati fotky z predchoziho behu zpet na puvodni misto."""
    log_path = dest / RUNS_DIR / f"beh_{run_id}.csv"
    if not log_path.exists():
        print(f"Log behu nenalezen: {log_path}")
        available = sorted((dest / RUNS_DIR).glob("beh_*.csv")) if (dest / RUNS_DIR).exists() else []
        if available:
            print("Dostupne behy:")
            for a in available:
                print(f"   {a.stem.replace('beh_', '')}")
        return 1

    restored = failed = skipped = 0
    with open(log_path, encoding="utf-8-sig") as f:
        rows = list(csv.DictReader(f))

    for row in rows:
        mode, src, dst = row["rezim"], Path(row["zdroj"]), Path(row["cil"])
        if not dst.exists():
            skipped += 1
            continue
        try:
            if mode == "move":
                src.parent.mkdir(parents=True, exist_ok=True)
                shutil.move(str(dst), str(src))
            else:
                dst.unlink()      # u kopirovani staci smazat kopii
            restored += 1
        except Exception as e:
            print(f"  [!] {dst.name}: {e}")
            failed += 1

    # Uklid prazdnych slozek, ktere po vraceni zbyly
    for folder in sorted(dest.rglob("*"), key=lambda p: len(p.parts), reverse=True):
        if folder.is_dir() and not any(folder.iterdir()) and folder.name != RUNS_DIR:
            try:
                folder.rmdir()
            except Exception:
                pass

    print(f"\nVraceno zpet: {restored}   preskoceno (uz neexistuje): {skipped}   chyb: {failed}")
    if restored:
        print("POZOR: evidence zpracovanych fotek zustava - pokud chcete fotky zpracovat")
        print(f"znovu, spustte trideni s prepinacem --reprocess.")
    return 0 if failed == 0 else 1


# ==========================================================================
# EXCEL REPORT
# ==========================================================================

HEADER_FILL = "1F4E5F"
WARN_FILL = "FFF2CC"
BAD_FILL = "F8CBAD"
OK_FILL = "E2EFDA"


def _style_sheet(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def write_excel_report(dest: Path, run_id: str, summary: dict, photos: list,
                       rooms_stats: list, anomalies: list, tech_map: dict):
    """Jeden Excel se vsemi listy + pripojeni radku do souhrnneho prehledu."""
    if not XLSX_OK:
        print("  [!] openpyxl neni nainstalovan - Excel report se preskakuje.")
        print("      Nainstalujte:  pip install openpyxl")
        return None

    wb = openpyxl.Workbook()

    # --- List 1: Souhrn ---
    ws = wb.active
    ws.title = "Souhrn"
    ws.append(["Polozka", "Hodnota"])
    for key, val in summary.items():
        ws.append([key, val])
    _style_sheet(ws, [42, 60])

    # --- List 2: Ke kontrole (to nejdulezitejsi - co je potreba rucne dodelat) ---
    ws = wb.create_sheet("Ke kontrole")
    ws.append(["Typ", "Mistnost / slozka", "Pocet fotek", "Popis", "Prvni fotka"])
    for a in anomalies:
        ws.append([a["typ"], a["slozka"], a["pocet"], a["popis"], a["prvni_fotka"]])
        fill = BAD_FILL if a["typ"].startswith("NEZNAM") else WARN_FILL
        for cell in ws[ws.max_row]:
            cell.fill = PatternFill("solid", fgColor=fill)
    if not anomalies:
        ws.append(["-", "-", 0, "Zadne anomalie, vse vypada v poradku.", "-"])
        for cell in ws[ws.max_row]:
            cell.fill = PatternFill("solid", fgColor=OK_FILL)
    _style_sheet(ws, [22, 26, 12, 62, 30])

    # --- List 3: Mistnosti ---
    ws = wb.create_sheet("Mistnosti")
    ws.append(["Budova", "Mistnost", "Technologie", "Pocet fotek", "Zdroj rozpoznani",
               "Skore shody", "Od", "Do", "Stav"])
    for r in rooms_stats:
        ws.append([r["budova"], r["mistnost"], r["technologie"], r["pocet"],
                   r["zdroj"], r["skore"], r["od"], r["do"], r["stav"]])
        if r["stav"] != "OK":
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill("solid", fgColor=WARN_FILL)
    _style_sheet(ws, [12, 16, 14, 12, 18, 12, 20, 20, 28])

    # --- List 4: Vsechny fotky ---
    ws = wb.create_sheet("Fotky")
    ws.append(["Poradi", "Soubor", "Cas poriz.", "Tablet?", "Mistnost", "Budova",
               "Technologie", "Zdroj", "Skore", "Kandidati z OCR", "Cilova cesta", "OCR (zkraceno)"])
    for p in photos:
        ws.append([p["poradi"], p["soubor"], p["cas"], "ANO" if p["tablet"] else "",
                   p["mistnost"], p["budova"], p["technologie"], p["zdroj"], p["skore"],
                   "; ".join(p["kandidati"])[:200], p["cil"], p["ocr"][:200]])
    _style_sheet(ws, [8, 26, 20, 9, 14, 10, 13, 14, 9, 34, 52, 60])

    # --- List 5: Technologie ---
    ws = wb.create_sheet("Technologie")
    ws.append(["Zkratka", "Nazev", "Pocet fotek"])
    tech_counts = {}
    for p in photos:
        tech_counts[p["technologie"]] = tech_counts.get(p["technologie"], 0) + 1
    for code, count in sorted(tech_counts.items(), key=lambda x: -x[1]):
        ws.append([code, tech_map.get(code.upper(), ""), count])
    _style_sheet(ws, [16, 40, 14])

    out_path = dest / f"report_{run_id}.xlsx"
    wb.save(out_path)

    _append_master_overview(dest, run_id, summary)
    return out_path


def _append_master_overview(dest: Path, run_id: str, summary: dict):
    """Prida radek do souhrnneho Excelu, ktery drzi historii vsech behu."""
    master = dest / MASTER_XLSX
    headers = ["Beh (run id)", "Datum a cas", "Technologie", "Fotek celkem",
               "Rozpoznanych mistnosti", "Neznamych serii", "Preskoceno (jiz zpracovano)",
               "Anomalii k reseni", "Vision volani", "Doba behu", "Rezim", "Report"]
    try:
        if master.exists():
            wb = openpyxl.load_workbook(master)
            ws = wb["Historie behu"] if "Historie behu" in wb.sheetnames else wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Historie behu"
            ws.append(headers)
            _style_sheet(ws, [20, 20, 14, 14, 20, 16, 24, 18, 14, 12, 12, 26])
        ws.append([
            run_id,
            summary.get("Datum a cas", ""),
            summary.get("Technologie", ""),
            summary.get("Fotek ke zpracovani", 0),
            summary.get("Rozpoznanych mistnosti", 0),
            summary.get("Neznamych serii", 0),
            summary.get("Preskoceno (jiz zpracovano drive)", 0),
            summary.get("Anomalii k rucni kontrole", 0),
            summary.get("Volani vision API", 0),
            summary.get("Doba behu", ""),
            summary.get("Rezim", ""),
            f"report_{run_id}.xlsx",
        ])
        wb.save(master)
    except PermissionError:
        print(f"  [!] {MASTER_XLSX} je otevreny v Excelu - souhrn se nepripsal. Zavrete ho a spustte znovu.")
    except Exception as e:
        print(f"  [!] Nepodarilo se aktualizovat {MASTER_XLSX}: {e}")


def write_csv_fallback(dest: Path, run_id: str, photos: list):
    """Zaloha pro pripad, ze openpyxl neni k dispozici."""
    path = dest / f"report_{run_id}.csv"
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(photos[0].keys()))
        w.writeheader()
        for p in photos:
            row = dict(p)
            row["kandidati"] = "; ".join(p["kandidati"])
            w.writerow(row)
    return path


# ==========================================================================
# REZIM: LEARN (zpetna vazba z rucnich oprav)
# ==========================================================================

def mode_learn(dest: Path, configs_dir: Path, index: RoomIndex):
    """Projde roztridene slozky a nauci se z nich nove nazvy mistnosti.

    Pouziti: kdyz rucne prejmenujete slozku 'Neznama_mistnost_3' na 'GS644',
    tento rezim ten kod najde a doplni ho do konfigurace budovy. Priste uz ho
    skript bude znat a rozpozna ho sam - presnost tak s kazdou rucni opravou roste."""
    print(f"Prohledavam {dest} ...")
    found = {}      # budova -> [nove mistnosti]
    unknown_folders = []

    for building_dir in sorted(p for p in dest.iterdir() if p.is_dir()):
        if building_dir.name.startswith("_"):
            continue
        building_code = building_dir.name.upper()
        for room_dir in sorted(p for p in building_dir.iterdir() if p.is_dir()):
            name = room_dir.name.strip().upper()
            if name.startswith(UNKNOWN_ROOM_PREFIX.upper()) or name.startswith("_"):
                unknown_folders.append(f"{building_dir.name}/{room_dir.name}")
                continue
            if name in index.room_to_building:
                continue
            found.setdefault(building_code, []).append(name)

    if unknown_folders:
        print(f"\nStale nevyresene slozky ({len(unknown_folders)}) - prejmenujte je na skutecny")
        print("kod mistnosti a spustte 'learn' znovu:")
        for u in unknown_folders[:30]:
            print(f"   {u}")
        if len(unknown_folders) > 30:
            print(f"   ... a dalsich {len(unknown_folders) - 30}")

    if not found:
        print("\nZadne nove mistnosti k naucení nenalezeny.")
        return 0

    total_added = 0
    for building_code, rooms in found.items():
        cfg = index.buildings.get(building_code)
        if cfg is None:
            # Budova zatim nema konfiguraci - zalozime ji
            new_path = configs_dir / f"{building_code}.json"
            new_path.write_text(json.dumps({
                "building_code": building_code,
                "room_code_pattern": DEFAULT_ROOM_PATTERN,
                "fuzzy_match_threshold": 0.8,
                "rooms": sorted(set(rooms)),
            }, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"\n[+] Zalozena nova konfigurace budovy: {new_path.name} "
                  f"({len(set(rooms))} mistnosti)")
            total_added += len(set(rooms))
            continue
        added = cfg.save_rooms(sorted(set(rooms)))
        if added:
            print(f"\n[+] {cfg.source_path.name}: pridano {len(added)} mistnosti")
            for a in added:
                print(f"      {a}")
            total_added += len(added)

    print(f"\nHotovo. Doplneno celkem {total_added} mistnosti do konfiguraci.")
    print("Zaloha puvodnich souboru je ulozena vedle nich s priponou .json.bak")
    return 0


# ==========================================================================
# REZIM: SORT (hlavni trideni)
# ==========================================================================

def collect_batches(args) -> list:
    """Vraci seznam davek ke zpracovani: (slozka, technologie).

    Rezim --inbox: kazda podslozka inboxu = jedna technologie (napr. VZT).
    Rezim --source: jedna slozka, technologie z --technology."""
    batches = []
    if args.inbox:
        inbox = Path(args.inbox)
        if not inbox.is_dir():
            print(f"Inbox neexistuje: {inbox}")
            sys.exit(1)
        subdirs = sorted(p for p in inbox.iterdir() if p.is_dir() and not p.name.startswith("_"))
        if subdirs:
            for sub in subdirs:
                batches.append((sub, sub.name.upper()))
        else:
            # Fotky lezi primo v inboxu bez podslozky technologie
            batches.append((inbox, args.technology.upper() if args.technology else ""))
    else:
        source = Path(args.source)
        if not source.is_dir():
            print(f"Zdrojova slozka neexistuje: {source}")
            sys.exit(1)
        batches.append((source, args.technology.upper() if args.technology else ""))
    return batches


def mode_sort(args, index: RoomIndex, tech_map: dict) -> int:
    dest = Path(args.dest)
    dest.mkdir(parents=True, exist_ok=True)
    run_id = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    started = time.time()

    ledger = Ledger(dest / LEDGER_FILE)
    vision_cache = VisionCache(dest / VISION_CACHE_FILE) if args.vision_fallback else None

    batches = collect_batches(args)
    print(f"Davky ke zpracovani: {len(batches)}")
    for folder, tech in batches:
        print(f"   {folder.name}  ->  technologie: {tech or '(neurcena)'}")

    # ---------- Sber souboru ----------
    all_items = []      # (path, technologie)
    skipped_known = 0
    for folder, tech in batches:
        files = sorted(p for p in folder.rglob("*") if p.suffix.lower() in IMAGE_EXTS)
        for f in files:
            if not args.reprocess:
                try:
                    if ledger.seen(quick_hash(f)):
                        skipped_known += 1
                        continue
                except Exception:
                    pass
            all_items.append((f, tech))

    if not all_items:
        print("\nZadne nove fotky ke zpracovani.")
        if skipped_known:
            print(f"({skipped_known} fotek uz bylo zpracovano drive - pro znovuzpracovani pouzijte --reprocess)")
        return 0

    print(f"\nFotek ke zpracovani: {len(all_items)}"
          + (f"   (preskoceno jiz zpracovanych: {skipped_known})" if skipped_known else ""))

    heic_count = sum(1 for p, _ in all_items if p.suffix.lower() in (".heic", ".heif"))
    if heic_count and not HEIC_OK:
        print(f"\n  [!] Nalezeno {heic_count} HEIC fotek, ale knihovna pillow-heif chybi.")
        print("      Nainstalujte:  pip install pillow-heif    (jinak budou preskoceny)")

    # ---------- FAZE 1: paralelni analyza ----------
    workers = args.workers or max(1, (os.cpu_count() or 4) - 1)
    print(f"\nAnalyza fotek ({workers} paralelnich procesu, orez displeje: "
          f"{'ano' if (CV2_OK and not args.no_screen_crop) else 'ne'})...")

    paths = [str(p) for p, _ in all_items]
    tech_by_path = {str(p): t for p, t in all_items}
    results = {}

    init_args = (
        args.tesseract, args.lang, index.pattern.pattern, index.all_rooms,
        index.room_to_building, index.fuzzy_threshold, CV2_OK and not args.no_screen_crop,
    )

    done = 0
    if workers == 1:
        _worker_init(*init_args)
        for path_str in paths:
            results[path_str] = _worker_analyze(path_str)
            done += 1
            if done % 25 == 0 or done == len(paths):
                print(f"   {done}/{len(paths)}")
    else:
        with ProcessPoolExecutor(max_workers=workers, initializer=_worker_init,
                                 initargs=init_args) as pool:
            for res in pool.map(_worker_analyze, paths, chunksize=4):
                results[res["path"]] = res
                done += 1
                if done % 25 == 0 or done == len(paths):
                    print(f"   {done}/{len(paths)}")

    errors = [r for r in results.values() if r["error"]]
    if errors:
        print(f"\n  [!] {len(errors)} fotek se nepodarilo zpracovat (viz Excel report).")

    # ---------- FAZE 2: sekvencni prirazeni do mistnosti ----------
    print("\nPrirazovani fotek do mistnosti (chronologicky)...")

    ordered = sorted(paths, key=lambda p: (results[p]["exif"] or "9999", Path(p).name))
    missing_exif = sum(1 for p in ordered if not results[p]["exif"])
    if missing_exif:
        print(f"   Pozor: {missing_exif} fotek nema cas poriz. -> radi se podle nazvu souboru.")

    photos = []
    series = []             # seznam serii: dict(budova, mistnost, technologie, fotky[], ...)
    current = None
    unknown_counter = 0
    vision_calls = 0
    has_room_db = bool(index.all_rooms)

    for i, path_str in enumerate(ordered, 1):
        res = results[path_str]
        path = Path(path_str)
        tech = tech_by_path.get(path_str, "")
        zdroj = ""

        if res["is_tablet"]:
            room = res["room"]
            building = res["building"]
            score = res["score"]

            if has_room_db:
                # Mame seznam skutecnych mistnosti -> rozhoduje shoda s nim.
                confident = (bool(room) and score >= STRONG_MATCH_SCORE
                             and res["hits"] >= args.min_confidence)
                fuzzy_ok = bool(room) and score >= index.fuzzy_threshold
                zdroj = "OCR (jista shoda)" if confident else ("OCR (podobnost)" if fuzzy_ok else "")
            else:
                # Bez seznamu mistnosti nelze skore spocitat - rozhoduje jen to,
                # jak konzistentne se kod objevil napric OCR variantami.
                confident = bool(room) and res["hits"] >= args.min_confidence
                fuzzy_ok = confident
                zdroj = "OCR (bez seznamu mistnosti)" if confident else ""

            # Vision fallback jen kdyz si lokalne nejsme jisti
            if not fuzzy_ok and args.vision_fallback and not args.dry_run:
                try:
                    key = quick_hash(path)
                except Exception:
                    key = path.name
                cached = vision_cache.get(key)
                if cached is None:
                    vision_calls += 1
                    cached = call_vision(path, index.all_rooms)
                    vision_cache.set(key, cached)
                    zdroj = "Vision API"
                else:
                    zdroj = "Vision (cache)"
                if cached:
                    matched, matched_building, matched_score = index.match(cached)
                    if matched and matched_score >= index.fuzzy_threshold:
                        room, building, score = matched, matched_building, matched_score
                    else:
                        room, building, score = cached, "", 0.0
                    fuzzy_ok = True

            if fuzzy_ok and room:
                building = building or UNKNOWN_BUILDING
                current = {
                    "budova": building, "mistnost": room, "technologie": tech,
                    "zdroj": zdroj, "skore": round(score, 3), "fotky": [],
                    "prvni_fotka": path.name,
                }
                series.append(current)
            else:
                unknown_counter += 1
                current = {
                    "budova": UNKNOWN_BUILDING,
                    "mistnost": f"{UNKNOWN_ROOM_PREFIX}_{unknown_counter}",
                    "technologie": tech, "zdroj": "nerozpoznano", "skore": 0.0,
                    "fotky": [], "prvni_fotka": path.name,
                }
                series.append(current)
                zdroj = zdroj or "nerozpoznano"

        if current is None:
            # Fotky pred prvni fotkou tabletu
            current = {"budova": UNKNOWN_BUILDING, "mistnost": BEFORE_FIRST_TABLET,
                       "technologie": tech, "zdroj": "pred prvnim tabletem", "skore": 0.0,
                       "fotky": [], "prvni_fotka": path.name}
            series.append(current)

        current["fotky"].append(path)

        rel = Path(current["budova"]) / current["mistnost"]
        if tech:
            rel = rel / tech
        photos.append({
            "poradi": i,
            "soubor": path.name,
            "cas": res["exif"],
            "tablet": res["is_tablet"],
            "mistnost": current["mistnost"],
            "budova": current["budova"],
            "technologie": tech,
            "zdroj": zdroj if res["is_tablet"] else "",
            "skore": round(res["score"], 3) if res["is_tablet"] else "",
            "kandidati": res["candidates"],
            "cil": str(rel / path.name),
            "ocr": res["ocr_text"].replace("\n", " | ") if res["is_tablet"] else "",
            "chyba": res["error"],
            "_src": path,
            "_rel": rel,
        })

    # ---------- FAZE 3: zapis souboru ----------
    moves = []
    written = 0
    write_errors = []

    if args.dry_run:
        print("\n[DRY RUN] Nic se nepresouva - jen se pocita, co by se stalo.")
    else:
        print(f"\n{'Presouvam' if args.move else 'Kopiruji'} fotky...")
        for p in photos:
            target_dir = dest / p["_rel"]
            try:
                target_dir.mkdir(parents=True, exist_ok=True)
                target = target_dir / p["_src"].name
                if target.exists():
                    target = target_dir / f"{p['_src'].stem}_dup{p['_src'].suffix}"
                src_hash = quick_hash(p["_src"])
                if args.move:
                    shutil.move(str(p["_src"]), str(target))
                else:
                    shutil.copy2(str(p["_src"]), str(target))
                moves.append((str(p["_src"]), str(target)))
                ledger.record(src_hash, str(p["_src"]), str(target), run_id)
                p["cil"] = str(target.relative_to(dest))
                written += 1
            except Exception as e:
                write_errors.append(f"{p['soubor']}: {e}")

        ledger.save()
        log_path = write_run_log(dest / RUNS_DIR, run_id, moves,
                                 "move" if args.move else "copy")
        print(f"Log behu (pro pripadne vraceni zpet): {log_path.name}")

    # ---------- FAZE 4: kontrola integrity ----------
    print("\nKontrola integrity...")
    integrity_msgs = []
    if args.dry_run:
        integrity_msgs.append("DRY RUN - kontrola integrity se neprovadi")
    else:
        if written != len(photos):
            integrity_msgs.append(
                f"CHYBA: ke zpracovani bylo {len(photos)} fotek, ale zapsano jen {written}")
        missing = [src for src, dst in moves if not Path(dst).exists()]
        if missing:
            integrity_msgs.append(f"CHYBA: {len(missing)} cilovych souboru neexistuje")
        if args.move:
            leftover = [src for src, dst in moves if Path(src).exists()]
            if leftover:
                integrity_msgs.append(
                    f"POZOR: {len(leftover)} zdrojovych souboru po presunu stale existuje")
        for err in write_errors:
            integrity_msgs.append(f"CHYBA zapisu: {err}")
        if not integrity_msgs:
            integrity_msgs.append(f"OK - vsech {written} fotek zapsano a overeno")
    for msg in integrity_msgs:
        print(f"   {msg}")

    # ---------- FAZE 5: anomalie ----------
    anomalies = []
    room_occurrences = {}
    for s in series:
        if s["mistnost"] not in (BEFORE_FIRST_TABLET,):
            key = (s["budova"], s["mistnost"])
            room_occurrences.setdefault(key, []).append(s)

    for s in series:
        count = len(s["fotky"])
        name = s["mistnost"]
        if name.startswith(UNKNOWN_ROOM_PREFIX):
            anomalies.append({
                "typ": "NEZNAMA MISTNOST", "slozka": name, "pocet": count,
                "popis": "Fotka tabletu nalezena, ale nazev mistnosti se nepodarilo precist. "
                         "Prejmenujte slozku na spravny kod a spustte rezim 'learn'.",
                "prvni_fotka": s["prvni_fotka"],
            })
        elif name == BEFORE_FIRST_TABLET:
            anomalies.append({
                "typ": "PRED PRVNIM TABLETEM", "slozka": name, "pocet": count,
                "popis": "Fotky pred prvni fotkou tabletu - chybi jim urceni mistnosti.",
                "prvni_fotka": s["prvni_fotka"],
            })
        elif count < EXPECTED_SERIES_MIN:
            anomalies.append({
                "typ": "MALO FOTEK", "slozka": name, "pocet": count,
                "popis": f"Serie ma jen {count} fotek (obvykle {EXPECTED_SERIES_MIN}-"
                         f"{EXPECTED_SERIES_MAX}). Mozna byla omylem rozpoznana fotka tabletu navic.",
                "prvni_fotka": s["prvni_fotka"],
            })
        elif count > EXPECTED_SERIES_MAX:
            anomalies.append({
                "typ": "HODNE FOTEK", "slozka": name, "pocet": count,
                "popis": f"Serie ma {count} fotek (obvykle {EXPECTED_SERIES_MIN}-"
                         f"{EXPECTED_SERIES_MAX}). Mozna byla prehlednuta fotka tabletu uprostred.",
                "prvni_fotka": s["prvni_fotka"],
            })
        if s["zdroj"] == "OCR (podobnost)":
            anomalies.append({
                "typ": "NEJISTA SHODA", "slozka": name, "pocet": count,
                "popis": f"Kod mistnosti byl doplnen podle podobnosti (skore {s['skore']}), "
                         f"ne presnou shodou. Doporucuji vizualne overit.",
                "prvni_fotka": s["prvni_fotka"],
            })

    for (building, room), occurrences in room_occurrences.items():
        if len(occurrences) > 1:
            anomalies.append({
                "typ": "ROZDELENA SERIE", "slozka": room,
                "pocet": sum(len(o["fotky"]) for o in occurrences),
                "popis": f"Mistnost se objevila {len(occurrences)}x v oddelenych blocich. "
                         f"Fotky jsou slouceny do jedne slozky - overte, ze je to spravne.",
                "prvni_fotka": occurrences[0]["prvni_fotka"],
            })

    # ---------- FAZE 6: report ----------
    rooms_stats = []
    for s in series:
        times = [results[str(f)]["exif"] for f in s["fotky"] if results.get(str(f))]
        times = [t for t in times if t]
        stav = "OK"
        for a in anomalies:
            if a["slozka"] == s["mistnost"]:
                stav = a["typ"]
                break
        rooms_stats.append({
            "budova": s["budova"], "mistnost": s["mistnost"], "technologie": s["technologie"],
            "pocet": len(s["fotky"]), "zdroj": s["zdroj"], "skore": s["skore"],
            "od": times[0] if times else "", "do": times[-1] if times else "",
            "stav": stav,
        })

    elapsed = time.time() - started
    recognized = sum(1 for s in series
                     if not s["mistnost"].startswith(UNKNOWN_ROOM_PREFIX)
                     and s["mistnost"] != BEFORE_FIRST_TABLET)
    summary = {
        "Beh (run id)": run_id,
        "Datum a cas": dt.datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
        "Rezim": "DRY RUN" if args.dry_run else ("presun" if args.move else "kopie"),
        "Technologie": ", ".join(sorted({t for _, t in batches if t})) or "(neurcena)",
        "Zdroj": args.inbox or args.source,
        "Cil": str(dest),
        "Fotek ke zpracovani": len(photos),
        "Preskoceno (jiz zpracovano drive)": skipped_known,
        "Fotek tabletu nalezeno": sum(1 for p in photos if p["tablet"]),
        "Rozpoznanych mistnosti": recognized,
        "Neznamych serii": sum(1 for s in series if s["mistnost"].startswith(UNKNOWN_ROOM_PREFIX)),
        "Anomalii k rucni kontrole": len(anomalies),
        "Volani vision API": vision_calls,
        "Chyb pri cteni fotek": len(errors),
        "Kontrola integrity": " | ".join(integrity_msgs),
        "Doba behu": f"{elapsed/60:.1f} min" if elapsed > 90 else f"{elapsed:.0f} s",
        "Paralelnich procesu": workers,
        "Orez displeje (OpenCV)": "ano" if (CV2_OK and not args.no_screen_crop) else "ne",
        "HEIC podpora": "ano" if HEIC_OK else "ne",
        "Znamych mistnosti v databazi": len(index.all_rooms),
    }

    report_path = write_excel_report(dest, run_id, summary, photos, rooms_stats,
                                     anomalies, tech_map)
    if report_path is None and photos:
        report_path = write_csv_fallback(dest, run_id, photos)

    # ---------- Vypis ----------
    print("\n" + "=" * 62)
    print(f"  HOTOVO   ({summary['Doba behu']})")
    print("=" * 62)
    print(f"  Fotek zpracovano:        {len(photos)}")
    print(f"  Rozpoznanych mistnosti:  {recognized}")
    print(f"  Neznamych serii:         {summary['Neznamych serii']}")
    print(f"  Anomalii k reseni:       {len(anomalies)}")
    if vision_calls:
        print(f"  Volani vision API:       {vision_calls}")
    if report_path:
        print(f"\n  Report:  {report_path.name}")
        print(f"  Souhrn:  {MASTER_XLSX}")
    if anomalies:
        print(f"\n  >> Otevrete report, list 'Ke kontrole' - je tam {len(anomalies)} polozek k rucnimu dodelani.")
    if summary["Neznamych serii"]:
        print("  >> Po rucnim prejmenovani slozek spustte:  --mode learn")
    print()
    return 0


# ==========================================================================
# CLI
# ==========================================================================

def main():
    ap = argparse.ArgumentParser(
        description="Trideni fotek z pasportizace do slozek podle mistnosti a technologie.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--mode", choices=["sort", "learn", "undo"], default="sort",
                    help="sort = trideni (vychozi), learn = uceni z rucnich oprav, undo = vraceni behu")
    ap.add_argument("--dest", required=True, help="Cilova slozka pasportu (koren).")
    ap.add_argument("--inbox", default=None,
                    help="Inbox slozka - kazda podslozka = jedna technologie (napr. VZT).")
    ap.add_argument("--source", default=None, help="Jedna zdrojova slozka (alternativa k --inbox).")
    ap.add_argument("--technology", default=None,
                    help="Zkratka technologie, kdyz se nepozna z nazvu slozky (napr. VZT).")
    ap.add_argument("--configs-dir", default=None,
                    help="Slozka s konfiguracemi budov (vychozi: buildings/ vedle skriptu).")
    ap.add_argument("--tesseract", default=None, help="Cesta k tesseract.exe.")
    ap.add_argument("--lang", default="ces+eng", help="Jazyky OCR (vychozi ces+eng).")
    ap.add_argument("--workers", type=int, default=None,
                    help="Pocet paralelnich procesu (vychozi: pocet jader minus 1).")
    ap.add_argument("--move", action="store_true",
                    help="Presunout fotky (vychozi je kopirovat = bezpecnejsi).")
    ap.add_argument("--dry-run", action="store_true",
                    help="Nic nepresouvat, jen spocitat a vypsat report. Doporuceno pro prvni beh.")
    ap.add_argument("--reprocess", action="store_true",
                    help="Zpracovat i fotky, ktere uz byly drive zpracovane.")
    ap.add_argument("--no-screen-crop", action="store_true",
                    help="Vypnout detekci a orez displeje tabletu.")
    ap.add_argument("--vision-fallback", action="store_true",
                    help="U necitelnych fotek zavolat Claude vision API (placene, ANTHROPIC_API_KEY).")
    ap.add_argument("--min-confidence", type=int, default=MIN_CONFIDENCE_HITS,
                    help="Min. pocet vyskytu kodu napric OCR variantami pro jistou shodu.")
    ap.add_argument("--run-id", default=None, help="ID behu pro --mode undo.")
    args = ap.parse_args()

    if args.tesseract:
        pytesseract.pytesseract.tesseract_cmd = args.tesseract

    dest = Path(args.dest)
    script_dir = Path(__file__).resolve().parent
    configs_dir = Path(args.configs_dir) if args.configs_dir else script_dir / "buildings"

    # --- undo ---
    if args.mode == "undo":
        if not args.run_id:
            print("Pro --mode undo je potreba --run-id (najdete v nazvu logu ve slozce _behy).")
            runs = sorted((dest / RUNS_DIR).glob("beh_*.csv")) if (dest / RUNS_DIR).exists() else []
            if runs:
                print("\nDostupne behy:")
                for r in runs[-10:]:
                    print(f"   {r.stem.replace('beh_', '')}")
            return 1
        return undo_run(dest, args.run_id)

    # --- nacteni konfiguraci budov ---
    if not configs_dir.exists():
        print(f"Slozka s konfiguracemi budov neexistuje: {configs_dir}")
        print("Vytvorte ji a vlozte do ni JSON konfigurace budov (viz G62.example.json).")
        return 1
    index = RoomIndex.load_dir(configs_dir)
    if index.buildings:
        print(f"Nacteno budov: {len(index.buildings)} "
              f"({', '.join(sorted(index.buildings))}), "
              f"mistnosti celkem: {len(index.all_rooms)}")
    else:
        print(f"  [!] Ve slozce {configs_dir} nejsou zadne konfigurace budov.")
        print("      Skript pobezi, ale bez seznamu mistnosti bude vyrazne mene presny.")

    # --- learn ---
    if args.mode == "learn":
        if not dest.is_dir():
            print(f"Cilova slozka neexistuje: {dest}")
            return 1
        return mode_learn(dest, configs_dir, index)

    # --- sort ---
    if not args.inbox and not args.source:
        print("Zadejte --inbox (doporuceno) nebo --source.")
        return 1
    if not HEIC_OK:
        print("  [i] pillow-heif neni nainstalovan - HEIC fotky z iPhonu nepujdou precist.")
    if not CV2_OK:
        print("  [i] OpenCV neni nainstalovan - preskoci se orez displeje tabletu (nizsi presnost).")
    if not XLSX_OK:
        print("  [i] openpyxl neni nainstalovan - misto Excelu se zapise jen CSV.")

    tech_map = load_technologies(configs_dir / "technologie.json")
    return mode_sort(args, index, tech_map)


if __name__ == "__main__":
    sys.exit(main())
